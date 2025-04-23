from fastapi import FastAPI, Request, Form, UploadFile, File, Depends, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from starlette.middleware.sessions import SessionMiddleware
from typing import List, Dict, Tuple
import openpyxl
import os
import shutil
import secrets
from collections import defaultdict

app = FastAPI()
app.add_middleware(SessionMiddleware, secret_key="your-secret-key")

templates = Jinja2Templates(directory="templates")
app.mount("/static", StaticFiles(directory="static"), name="static")

FILE_PATHS = {
    "product": "product_data.xlsx",
    "material": "material_data.xlsx"
}
LOG_PATHS = {
    "product": "product_log.xlsx",
    "material": "material_log.xlsx"
}
TEMPLATE_FILE = "upload_template.xlsx"
PRODUCT_TEMPLATE_FILE = "product_upload_template.xlsx"

USERNAME = "admin"
PASSWORD = "maumiga123"

def init_excel():
    for key in FILE_PATHS:
        if not os.path.exists(FILE_PATHS[key]):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["제품명", "재고수량"])
            wb.save(FILE_PATHS[key])
        if not os.path.exists(LOG_PATHS[key]):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["일자", "입출고", "사유", "제품명", "수량"])
            wb.save(LOG_PATHS[key])
        if not os.path.exists(TEMPLATE_FILE):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "입출고양식"
            ws.append(["일자", "입출고", "사유", "제품명", "수량", "소비기한"])  # 소비기한 추가
            ws.append(["2025-04-28", "입고", "쿠팡", "콩쑥개떡", 10, "2025-06-30"])  # 예시 데이터 포함
            wb.save(TEMPLATE_FILE)
    if not os.path.exists(PRODUCT_TEMPLATE_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "신규제품양식"
        ws.append(["제품명", "수량"])
        ws.append(["감자절편", 25])
        wb.save(PRODUCT_TEMPLATE_FILE)

def get_current_user(request: Request):
    user = request.session.get("user")
    if not user:
        raise HTTPException(status_code=302, headers={"Location": "/login"})
    return user

@app.get("/login", response_class=HTMLResponse)
async def login_form(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})

@app.post("/login")
async def login(request: Request, username: str = Form(...), password: str = Form(...)):
    if username == USERNAME and password == PASSWORD:
        request.session["user"] = username
        return RedirectResponse(url="/", status_code=302)
    return templates.TemplateResponse("login.html", {"request": request, "error": "아이디 또는 비밀번호가 올바르지 않습니다."})

@app.get("/logout")
async def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login")

@app.get("/", response_class=HTMLResponse)
async def home(request: Request, username: str = Depends(get_current_user)):
    return templates.TemplateResponse("home.html", {"request": request, "username": username})

@app.get("/{category}", response_class=HTMLResponse)
async def view_category(request: Request, category: str, start: str = "", end: str = "", msg: str = "", username: str = Depends(get_current_user)):
    if category not in FILE_PATHS:
        return HTMLResponse("❌ 수불부 종류 오류", status_code=404)

    wb = openpyxl.load_workbook(FILE_PATHS[category])
    ws = wb.active
    products = list(ws.iter_rows(min_row=2, values_only=True))

    log_wb = openpyxl.load_workbook(LOG_PATHS[category])
    log_ws = log_wb.active
    logs = list(log_ws.iter_rows(min_row=2, values_only=True))

    def in_range(date_str):
        if not start and not end:
            return True
        if start and end:
            return start <= date_str <= end
        elif start:
            return date_str >= start
        elif end:
            return date_str <= end
        return True

    incoming_logs = [row for row in logs if str(row[1]).strip() == "입고" and in_range(str(row[0]))]
    outgoing_logs = [row for row in logs if str(row[1]).strip() == "출고" and in_range(str(row[0]))]

    def group_logs(log_list: List[Tuple]) -> Dict[str, List[Tuple]]:
        grouped = defaultdict(list)
        for row in log_list:
            key = f"{row[0]}|{row[2]}"  # 일자|사유
            grouped[key].append(row)
        return dict(grouped)

    grouped_incoming = group_logs(incoming_logs)
    grouped_outgoing = group_logs(outgoing_logs)

    return templates.TemplateResponse(f"{category}.html", {
        "request": request,
        "category": category,
        "products": products,
        "grouped_incoming": grouped_incoming,
        "grouped_outgoing": grouped_outgoing,
        "msg": msg,
        "start": start,
        "end": end,
        "username": username
    })

@app.post("/{category}/add")
async def add_product(category: str, name: str = Form(...), quantity: int = Form(...)):
    wb = openpyxl.load_workbook(FILE_PATHS[category])
    ws = wb.active
    ws.append([name, quantity])
    wb.save(FILE_PATHS[category])

    log_wb = openpyxl.load_workbook(LOG_PATHS[category])
    log_ws = log_wb.active
    log_ws.append(["신규등록", "입고", "신규 제품 등록", name, quantity])
    log_wb.save(LOG_PATHS[category])

    return RedirectResponse(f"/{category}", status_code=303)

@app.post("/{category}/delete")
async def delete_product(category: str, name: str = Form(...)):
    wb = openpyxl.load_workbook(FILE_PATHS[category])
    ws = wb.active
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value == name:
            ws.delete_rows(idx)
            break
    wb.save(FILE_PATHS[category])
    return RedirectResponse(f"/{category}", status_code=303)

@app.post("/{category}/record")
async def record_entry(category: str, date: str = Form(...), action: str = Form(...), reason: str = Form(...), name: str = Form(...), quantity: int = Form(...), expire: str = Form("")):
    wb = openpyxl.load_workbook(FILE_PATHS[category])
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == name:
            current = row[1].value or 0
            row[1].value = current + quantity if action == "입고" else max(0, current - quantity)
            break
    wb.save(FILE_PATHS[category])

    log_wb = openpyxl.load_workbook(LOG_PATHS[category])
    log_ws = log_wb.active
    log_ws.append([date, action, reason, name, quantity, expire])
    log_wb.save(LOG_PATHS[category])

    return RedirectResponse(f"/{category}", status_code=303)

@app.post("/{category}/upload")
async def upload_excel(category: str, file: UploadFile = File(...)):
    temp_path = f"temp_{category}.xlsx"
    with open(temp_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    wb = openpyxl.load_workbook(temp_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 6:
            date, action, reason, name, quantity = row
            expire = ""
        else:
            date, action, reason, name, quantity, expire = row

        data_wb = openpyxl.load_workbook(FILE_PATHS[category])
        data_ws = data_wb.active
        for data_row in data_ws.iter_rows(min_row=2):
            if data_row[0].value == name:
                current = data_row[1].value or 0
                data_row[1].value = current + quantity if action == "입고" else max(0, current - quantity)
                break
        data_wb.save(FILE_PATHS[category])

        log_wb = openpyxl.load_workbook(LOG_PATHS[category])
        log_ws = log_wb.active
        log_ws.append([date, action, reason, name, quantity, expire])
        log_wb.save(LOG_PATHS[category])

    os.remove(temp_path)
    return RedirectResponse(f"/{category}?msg=✅ 엑셀 업로드 완료", status_code=303)

@app.post("/{category}/upload-new")
async def upload_new_products(category: str, file: UploadFile = File(...)):
    temp_path = f"temp_new_{category}.xlsx"
    with open(temp_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    wb = openpyxl.load_workbook(temp_path)
    ws = wb.active
    data_wb = openpyxl.load_workbook(FILE_PATHS[category])
    data_ws = data_wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        name, quantity = row
        data_ws.append([name, quantity])

    data_wb.save(FILE_PATHS[category])
    os.remove(temp_path)
    return RedirectResponse(f"/{category}?msg=✅ 신규 제품 업로드 완료", status_code=303)

@app.post("/{category}/delete-log")
async def delete_log(category: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["일자", "입출고", "사유", "제품명", "수량"])
    wb.save(LOG_PATHS[category])
    return RedirectResponse(f"/{category}?msg=🧹 이력 삭제 완료", status_code=303)

@app.post("/{category}/delete-selected-log")
async def delete_selected_log(category: str, logs: List[str] = Form(...)):
    wb = openpyxl.load_workbook(LOG_PATHS[category])
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header, data = rows[0], rows[1:]

    targets = set()
    for log in logs:
        try:
            parts = log.split("|")
            if len(parts) != 3:
                continue  # 잘못된 형식 무시
            date_part, action, name = parts
            targets.add((date_part.strip(), action, name))
        except Exception:
            continue  # split 실패해도 무시하고 계속

    data_wb = openpyxl.load_workbook(FILE_PATHS[category])
    data_ws = data_wb.active

    filtered = []
    for row in data:
        key = (str(row[0]), str(row[1]), str(row[3]))
        if key in targets:
            for data_row in data_ws.iter_rows(min_row=2):
                if data_row[0].value == row[3]:
                    current = data_row[1].value or 0
                    if row[1] == "입고":
                        data_row[1].value = max(0, current - row[4])
                    elif row[1] == "출고":
                        data_row[1].value = current + row[4]
        else:
            filtered.append(row)

    data_wb.save(FILE_PATHS[category])

    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.append(header)
    for row in filtered:
        new_ws.append(row)
    new_wb.save(LOG_PATHS[category])

    return RedirectResponse(f"/{category}", status_code=303)

@app.get("/download/template")
async def download_template():
    return FileResponse(TEMPLATE_FILE, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="입출고_업로드_양식.xlsx")

@app.get("/download/product-template")
async def download_product_template():
    return FileResponse(PRODUCT_TEMPLATE_FILE, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="신규제품_업로드_양식.xlsx")

init_excel()

from starlette.middleware.sessions import SessionMiddleware
app.add_middleware(SessionMiddleware, secret_key="your-secret-key")

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))  # Render에서 제공하는 포트
    uvicorn.run("app:app", host="0.0.0.0", port=port)















